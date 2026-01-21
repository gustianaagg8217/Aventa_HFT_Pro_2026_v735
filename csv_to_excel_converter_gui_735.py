import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
import threading
from datetime import datetime

class CSVToExcelConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("CSV to Excel Converter Pro")
        self.root.geometry("700x600")
        self.root.resizable(False, False)
        
        # Apply dark theme
        self.root.configure(bg='#1a1e3a')
        
        # Variables
        self.csv_file_var = tk.StringVar()
        self.output_folder_var = tk.StringVar(value="Backtest")
        self.is_converting = False
        
        # Configure styles
        self.setup_styles()
        self.setup_ui()
    
    
    def setup_styles(self):
        """Setup dark theme styles to match main GUI"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Dark theme colors
        bg_color = '#1a1e3a'
        fg_color = '#e0e0e0'
        accent_color = '#00e676'
        warning_color = '#ffd600'
        error_color = '#ff1744'
        
        # Configure colors for ttk widgets
        style.configure('TFrame', background=bg_color)
        style.configure('TLabel', background=bg_color, foreground=fg_color)
        style.configure('TLabelframe', background=bg_color, foreground=fg_color)
        style.configure('TLabelframe.Label', background=bg_color, foreground=fg_color)
        style.configure('TButton', background=bg_color, foreground=fg_color, borderwidth=1)
        style.map('TButton', 
                  background=[('active', '#2a3f5f')],
                  foreground=[('active', accent_color)])
        style.configure('TEntry', fieldbackground='#0f1219', foreground=fg_color, borderwidth=1)
        style.configure('TCheckbutton', background=bg_color, foreground=fg_color)
        
    def setup_ui(self):
        """Setup UI components"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # === CSV FILE SELECTION ===
        file_frame = ttk.LabelFrame(main_frame, text="📁 CSV File Selection", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(file_frame, text="CSV File:", font=("Segoe UI", 10)).pack(anchor=tk.W, pady=5)
        
        file_input_frame = ttk.Frame(file_frame)
        file_input_frame.pack(fill=tk.X, pady=5)
        
        ttk.Entry(file_input_frame, textvariable=self.csv_file_var, width=50).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(file_input_frame, text="🔍 Browse", command=self.browse_csv, width=12).pack(side=tk.LEFT)
        
        # === OUTPUT FOLDER ===
        output_frame = ttk.LabelFrame(main_frame, text="📂 Output Folder", padding="10")
        output_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(output_frame, text="Output Folder:", font=("Segoe UI", 10)).pack(anchor=tk.W, pady=5)
        
        output_input_frame = ttk.Frame(output_frame)
        output_input_frame.pack(fill=tk.X, pady=5)
        
        ttk.Entry(output_input_frame, textvariable=self.output_folder_var, width=50).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        ttk.Button(output_input_frame, text="🔍 Browse", command=self.browse_output, width=12).pack(side=tk.LEFT)
        
        # === OPTIONS ===
        options_frame = ttk.LabelFrame(main_frame, text="⚙️ Options", padding="10")
        options_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(options_frame, text="Output Format:", font=("Segoe UI", 9)).pack(anchor=tk.W, pady=5)
        ttk.Label(options_frame, text="✓ Excel with colored profit/loss", font=("Segoe UI", 9), foreground="#2ecc71").pack(anchor=tk.W, padx=20)
        ttk.Label(options_frame, text="✓ Separate sheets: ALL_TRADES, BUY_TRADES, SELL_TRADES, SUMMARY", font=("Segoe UI", 9), foreground="#2ecc71").pack(anchor=tk.W, padx=20)
        
        # === CONTROL BUTTONS ===
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.convert_btn = ttk.Button(button_frame, text="🚀 Convert to Excel", command=self.convert_file, width=30)
        self.convert_btn.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="🗑️ Clear", command=self.clear_fields, width=15).pack(side=tk.LEFT, padx=5)
        
        # === LOG AREA ===
        log_frame = ttk.LabelFrame(main_frame, text="📝 Conversion Log", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.log_text = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=15, 
                                                  bg='#0f1219', fg='#e0e0e0', font=("Courier", 9),
                                                  insertbackground='#00e676')
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        self.log_text.tag_config("INFO", foreground="#00e676")
        self.log_text.tag_config("SUCCESS", foreground="#00b0ff")
        self.log_text.tag_config("ERROR", foreground="#ff1744")
        self.log_text.tag_config("WARNING", foreground="#ffd600")
        
        # === STATUS BAR ===
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, font=("Segoe UI", 9))
        status_bar.pack(fill=tk.X)
    
    def browse_csv(self):
        """Browse for CSV file"""
        filename = filedialog.askopenfilename(
            title="Select CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialdir="Backtest"
        )
        if filename:
            self.csv_file_var.set(filename)
            self.add_log(f"Selected file: {filename}", "INFO")
    
    def browse_output(self):
        """Browse for output folder"""
        folder = filedialog.askdirectory(title="Select Output Folder", initialdir="Backtest")
        if folder:
            self.output_folder_var.set(folder)
            self.add_log(f"Output folder: {folder}", "INFO")
    
    def clear_fields(self):
        """Clear all fields"""
        self.csv_file_var.set("")
        self.output_folder_var.set("Backtest")
        self.log_text.delete(1.0, tk.END)
        self.status_var.set("Ready")
        self.add_log("Fields cleared", "INFO")
    
    def add_log(self, message, level="INFO"):
        """Add message to log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        self.log_text.insert(tk.END, log_entry, level)
        self.log_text.see(tk.END)
        self.root.update()
    
    def convert_file(self):
        """Convert CSV to Excel in background thread"""
        if not self.csv_file_var.get():
            messagebox.showerror("Error", "Please select a CSV file first!")
            return
        
        if not os.path.exists(self.csv_file_var.get()):
            messagebox.showerror("Error", f"File not found: {self.csv_file_var.get()}")
            return
        
        # Create output folder if not exists
        output_folder = self.output_folder_var.get()
        if not os.path.exists(output_folder):
            try:
                os.makedirs(output_folder)
                self.add_log(f"Created output folder: {output_folder}", "INFO")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create output folder: {e}")
                return
        
        # Disable convert button and start conversion in thread
        self.convert_btn.config(state=tk.DISABLED)
        self.is_converting = True
        self.status_var.set("Converting...")
        
        thread = threading.Thread(target=self._do_conversion, daemon=True)
        thread.start()
    
    def _do_conversion(self):
        """Actual conversion logic (runs in background thread)"""
        try:
            csv_file = self.csv_file_var.get()
            output_folder = self.output_folder_var.get()
            
            self.add_log("=" * 60, "INFO")
            self.add_log("🚀 Starting conversion...", "INFO")
            self.add_log(f"Input: {csv_file}", "INFO")
            self.add_log(f"Output folder: {output_folder}", "INFO")
            
            # Load CSV
            self.add_log("📖 Loading CSV file...", "INFO")
            df = pd.read_csv(csv_file, skiprows=2)
            self.add_log(f"✓ Loaded {len(df)} trades", "SUCCESS")
            
            # Get initial balance
            self.add_log("💰 Extracting initial balance...", "INFO")
            initial_balance = 0.0
            try:
                with open(csv_file, 'r') as f:
                    first_line = f.readline().strip()
                    if 'Initial Balance' in first_line:
                        parts = first_line.split(',')
                        if len(parts) > 1:
                            initial_balance = float(parts[1].strip())
            except Exception as e:
                self.add_log(f"⚠️ Could not extract initial balance: {e}", "WARNING")
            
            self.add_log(f"✓ Initial Balance: ${initial_balance:.2f}", "SUCCESS")
            
            # Calculate metrics
            self.add_log("📊 Calculating metrics...", "INFO")
            
            COL_PROFIT = "Profit"
            COL_TYPE = "Type"
            
            buy_df = df[df[COL_TYPE].str.upper() == "BUY"]
            sell_df = df[df[COL_TYPE].str.upper() == "SELL"]
            
            total_trades = len(df)
            total_profit = df[COL_PROFIT].sum()
            
            win_trades = df[df[COL_PROFIT] > 0]
            loss_trades = df[df[COL_PROFIT] < 0]
            
            win_count = len(win_trades)
            loss_count = len(loss_trades)
            winrate = (win_count / total_trades) * 100 if total_trades else 0
            
            avg_profit = df[COL_PROFIT].mean()
            max_profit = df[COL_PROFIT].max()
            max_loss = df[COL_PROFIT].min()
            
            gross_profit = win_trades[COL_PROFIT].sum()
            gross_loss = abs(loss_trades[COL_PROFIT].sum())
            profit_factor = gross_profit / gross_loss if gross_loss != 0 else 0
            
            self.add_log(f"✓ Total Trades: {total_trades}", "INFO")
            self.add_log(f"✓ Win Rate: {winrate:.1f}%", "INFO")
            self.add_log(f"✓ Total Profit: ${total_profit:.2f}", "INFO")
            
            # Generate output filename
            self.add_log("📝 Generating output filename...", "INFO")
            input_basename = os.path.basename(os.path.splitext(csv_file)[0])
            output_file = os.path.join(output_folder, input_basename + "_report.xlsx")
            
            # Export to Excel
            self.add_log("📤 Exporting to Excel...", "INFO")
            
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="ALL_TRADES", index=False)
                self.add_log(f"  ✓ Exported ALL_TRADES sheet ({len(df)} rows)", "INFO")
                
                buy_df.to_excel(writer, sheet_name="BUY_TRADES", index=False)
                self.add_log(f"  ✓ Exported BUY_TRADES sheet ({len(buy_df)} rows)", "INFO")
                
                sell_df.to_excel(writer, sheet_name="SELL_TRADES", index=False)
                self.add_log(f"  ✓ Exported SELL_TRADES sheet ({len(sell_df)} rows)", "INFO")
                
                summary_df = pd.DataFrame({
                    "Metric": [
                        "Initial Balance ($)",
                        "Total Trades",
                        "Total Profit",
                        "Win Trades",
                        "Loss Trades",
                        "Winrate (%)",
                        "Average Profit",
                        "Max Profit",
                        "Max Loss",
                        "Profit Factor"
                    ],
                    "Value": [
                        round(initial_balance, 2),
                        total_trades,
                        round(total_profit, 2),
                        win_count,
                        loss_count,
                        round(winrate, 2),
                        round(avg_profit, 2),
                        round(max_profit, 2),
                        round(max_loss, 2),
                        round(profit_factor, 2)
                    ]
                })
                
                summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)
                self.add_log(f"  ✓ Exported SUMMARY sheet", "INFO")
            
            # Format Excel
            self.add_log("🎨 Formatting Excel...", "INFO")
            
            wb = load_workbook(output_file)
            
            def format_profit_column(ws):
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"{profit_col_letter}{row}"]
                    if cell.value is not None:
                        if cell.value > 0:
                            cell.font = Font(color="008000", bold=True)  # Green
                        elif cell.value < 0:
                            cell.font = Font(color="FF0000", bold=True)  # Red
            
            for sheet_name in ["ALL_TRADES", "BUY_TRADES", "SELL_TRADES"]:
                ws = wb[sheet_name]
                header = [cell.value for cell in ws[1]]
                if COL_PROFIT in header:
                    profit_col_letter = chr(header.index(COL_PROFIT) + 65)
                    format_profit_column(ws)
            
            # Bold header summary
            summary_ws = wb["SUMMARY"]
            for cell in summary_ws[1]:
                cell.font = Font(bold=True)
            
            wb.save(output_file)
            self.add_log(f"  ✓ Excel formatting applied", "INFO")
            
            self.add_log("=" * 60, "SUCCESS")
            self.add_log(f"✅ Conversion completed successfully!", "SUCCESS")
            self.add_log(f"📊 Output file: {output_file}", "SUCCESS")
            self.add_log("=" * 60, "SUCCESS")
            
            self.status_var.set(f"✓ Completed - Output: {output_file}")
            messagebox.showinfo("Success", f"Conversion completed!\n\nOutput: {output_file}")
            
        except Exception as e:
            error_msg = f"❌ Error: {str(e)}"
            self.add_log(error_msg, "ERROR")
            self.add_log("=" * 60, "ERROR")
            self.status_var.set("Error - Check logs")
            messagebox.showerror("Error", f"Conversion failed:\n\n{error_msg}")
        
        finally:
            self.is_converting = False
            self.root.after(0, lambda: self.convert_btn.config(state=tk.NORMAL))


def main():
    root = tk.Tk()
    app = CSVToExcelConverter(root)
    root.mainloop()


if __name__ == "__main__":
    main()
