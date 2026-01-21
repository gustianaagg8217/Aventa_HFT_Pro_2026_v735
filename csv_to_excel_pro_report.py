import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# =============================
# CONFIG
# =============================
CSV_FILE = "Backtest/trades_20260120_211634.csv"
OUTPUT_FILE = os.path.join("Backtest", os.path.basename(os.path.splitext(CSV_FILE)[0]) + "_convert.xlsx")

COL_PROFIT = "Profit"
COL_TYPE = "Type"     # BUY / SELL

# =============================
# LOAD DATA
# =============================
# Skip 2 baris pertama (Initial Balance dan empty row) dari format export GUI
df = pd.read_csv(CSV_FILE, skiprows=2)

# Get initial balance dari baris pertama CSV (sebelum skiprows)
initial_balance = 0.0
try:
    with open(CSV_FILE, 'r') as f:
        first_line = f.readline().strip()
        if 'Initial Balance' in first_line:
            parts = first_line.split(',')
            if len(parts) > 1:
                initial_balance = float(parts[1].strip())
except:
    initial_balance = 0.0

buy_df = df[df[COL_TYPE].str.upper() == "BUY"]
sell_df = df[df[COL_TYPE].str.upper() == "SELL"]

# =============================
# METRICS
# =============================
total_trades = len(df)
total_profit = df[COL_PROFIT].sum()

win_trades = df[df[COL_PROFIT] > 0]
loss_trades = df[df[COL_PROFIT] < 0]

win_count = len(win_trades)
loss_count = len(loss_trades)
winrate = (win_count / total_trades) * 100 if total_trades else 0

avg_profit = df[COL_PROFIT].mean()
max_profit = df[COL_PROFIT].max()
max_loss = df[COL_PROFIT].min()

gross_profit = win_trades[COL_PROFIT].sum()
gross_loss = abs(loss_trades[COL_PROFIT].sum())
profit_factor = gross_profit / gross_loss if gross_loss != 0 else 0

# =============================
# EXPORT EXCEL
# =============================
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="ALL_TRADES", index=False)
    buy_df.to_excel(writer, sheet_name="BUY_TRADES", index=False)
    sell_df.to_excel(writer, sheet_name="SELL_TRADES", index=False)

    summary_df = pd.DataFrame({
        "Metric": [
            "Initial Balance ($)",
            "Total Trades",
            "Total Profit",
            "Win Trades",
            "Loss Trades",
            "Winrate (%)",
            "Average Profit",
            "Max Profit",
            "Max Loss",
            "Profit Factor"
        ],
        "Value": [
            round(initial_balance, 2),
            total_trades,
            round(total_profit, 2),
            win_count,
            loss_count,
            round(winrate, 2),
            round(avg_profit, 2),
            round(max_profit, 2),
            round(max_loss, 2),
            round(profit_factor, 2)
        ]
    })

    summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)

# =============================
# FORMAT EXCEL
# =============================
wb = load_workbook(OUTPUT_FILE)

def format_profit_column(ws):
    for row in range(2, ws.max_row + 1):
        cell = ws[f"{profit_col_letter}{row}"]
        if cell.value is not None:
            if cell.value > 0:
                cell.font = Font(color="008000")
            elif cell.value < 0:
                cell.font = Font(color="FF0000")

for sheet_name in ["ALL_TRADES", "BUY_TRADES", "SELL_TRADES"]:
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    profit_col_letter = chr(header.index(COL_PROFIT) + 65)
    format_profit_column(ws)

# bold header summary
summary_ws = wb["SUMMARY"]
for cell in summary_ws[1]:
    cell.font = Font(bold=True)

wb.save(OUTPUT_FILE)

print("Selesai. Report profesional dibuat:", OUTPUT_FILE)
