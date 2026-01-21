import pandas as pd
from openpyxl import load_workbook

# =============================
# CONFIG
# =============================
CSV_FILE = "Backtest/trades_20260120_194856.csv"
OUTPUT_FILE = "Backtest/trades_summary.xlsx"
PROFIT_COLUMN = "Profit"   # sesuai file CSV kamu

# =============================
# LOAD CSV
# =============================
df = pd.read_csv(CSV_FILE)

# =============================
# HITUNG SUMMARY
# =============================
total_trades = len(df)
total_profit = df[PROFIT_COLUMN].sum()
avg_profit = df[PROFIT_COLUMN].mean()
max_profit = df[PROFIT_COLUMN].max()
min_profit = df[PROFIT_COLUMN].min()

# =============================
# EXPORT KE EXCEL
# =============================
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    # data CSV mulai dari row 9
    df.to_excel(writer, index=False, startrow=8, sheet_name="Trades")

# =============================
# EDIT HEADER SUMMARY
# =============================
wb = load_workbook(OUTPUT_FILE)
ws = wb["Trades"]

ws["A1"] = "SUMMARY"
ws["A2"] = "Total Trades"
ws["B2"] = total_trades

ws["A3"] = "Total Profit"
ws["B3"] = round(total_profit, 2)

ws["A4"] = "Average Profit"
ws["B4"] = round(avg_profit, 2)

ws["A5"] = "Max Profit"
ws["B5"] = round(max_profit, 2)

ws["A6"] = "Min Profit"
ws["B6"] = round(min_profit, 2)

# sedikit kosmetik biar enak dibaca
ws["A1"].font = ws["A2"].font.copy(bold=True)

wb.save(OUTPUT_FILE)

print("Sukses. File Excel dibuat:", OUTPUT_FILE)
